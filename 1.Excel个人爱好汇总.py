from openpyxl import load_workbook

wb = load_workbook('每个人的爱好.xlsx')
ws = wb.worksheets[0]
for index, row in enumerate(ws.rows):
    if index == 0:
        titles = tuple(map(lambda cell: cell.value, row))[1:]
        lastCol = len(titles)+2
        ws.cell(row=index+1, column=lastCol, value='所有爱好')
    else:
        values  = tuple(map(lambda cell: cell.value, row))[1:]
        result = '，'.join((titles[i] for i, v in enumerate(values) if v=='是'))
        ws.cell(row=index+1, column=lastCol, value=result)

wb.save('每个人的爱好汇总.xlsx')
